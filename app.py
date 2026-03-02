"""
Discord 频道导出工具 - Flask后端
支持论坛频道和普通频道，使用后台线程避免超时
"""

from flask import Flask, render_template, request, jsonify, send_file
import requests
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import re
import threading
import uuid

app = Flask(__name__)

# 默认Bot Token（从环境变量读取）
BOT_TOKEN = os.environ.get("DISCORD_BOT_TOKEN", "")

API_BASE = "https://discord.com/api/v9"

# 任务状态管理
export_lock = threading.Lock()
task_status = {
    "is_running": False,
    "task_id": None,
    "progress": "",
    "result": None,
    "error": None,
    "filename": None
}


def get_headers():
    return {"Authorization": f"Bot {BOT_TOKEN}"}


def snowflake_to_datetime(snowflake_id):
    timestamp_ms = (int(snowflake_id) >> 22) + 1420070400000
    return datetime.fromtimestamp(timestamp_ms / 1000)


def api_get(endpoint, params=None):
    url = f"{API_BASE}{endpoint}"
    while True:
        r = requests.get(url, headers=get_headers(), params=params)
        if r.status_code == 429:
            retry = r.json().get("retry_after", 1)
            time.sleep(retry)
            continue
        return r


def parse_discord_url(url):
    match = re.search(r"/channels/(\d+)/(\d+)", url)
    if match:
        return match.group(1), match.group(2)
    return None, None


def get_channel_info(channel_id):
    r = api_get(f"/channels/{channel_id}")
    if r.status_code == 200:
        return r.json()
    return None


def get_all_threads(forum_id):
    threads = []
    r = api_get(f"/channels/{forum_id}/threads/active")
    if r.status_code == 200:
        threads.extend(r.json().get("threads", []))

    params = {"limit": 100}
    while True:
        r = api_get(f"/channels/{forum_id}/threads/archived/public", params)
        if r.status_code != 200:
            break
        data = r.json()
        archived = data.get("threads", [])
        if not archived:
            break
        threads.extend(archived)
        if not data.get("has_more"):
            break
        params["before"] = archived[-1]["thread_metadata"]["archive_timestamp"]
    return threads


def get_channel_messages(channel_id, date_from=None, date_to=None):
    messages = []
    params = {"limit": 100}
    while True:
        r = api_get(f"/channels/{channel_id}/messages", params)
        if r.status_code != 200:
            break
        batch = r.json()
        if not batch:
            break
        for msg in batch:
            msg_time = snowflake_to_datetime(msg["id"])
            if date_from and msg_time < date_from:
                continue
            if date_to and msg_time > date_to:
                continue
            messages.append(msg)
        if len(batch) < 100:
            break
        params["before"] = batch[-1]["id"]
        time.sleep(0.3)
    return messages


def export_to_excel(threads_data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "消息导出"
    headers = ["频道/帖子", "创建时间", "消息作者", "消息时间", "消息内容", "附件", "消息链接"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
    row = 2
    for thread in threads_data:
        for msg in thread["messages"]:
            ws.cell(row=row, column=1, value=thread["name"])
            ws.cell(row=row, column=2, value=thread["created"])
            ws.cell(row=row, column=3, value=msg["author"])
            ws.cell(row=row, column=4, value=msg["time"])
            ws.cell(row=row, column=5, value=msg["content"])
            ws.cell(row=row, column=6, value=msg["attachments"])
            ws.cell(row=row, column=7, value=msg["link"])
            row += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 60
    wb.save(filename)
    return row - 1


def export_to_txt(threads_data, filename):
    with open(filename, "w", encoding="utf-8") as f:
        for thread in threads_data:
            f.write(f"{'='*60}\n频道/帖子: {thread['name']}\n创建时间: {thread['created']}\n{'='*60}\n\n")
            for msg in thread["messages"]:
                f.write(f"[{msg['time']}] {msg['author']}:\n{msg['content']}\n")
                if msg["attachments"]:
                    f.write(f"附件: {msg['attachments']}\n")
                f.write(f"链接: {msg['link']}\n\n")


def export_to_html(threads_data, filename):
    html = """<!DOCTYPE html><html><head><meta charset="utf-8"><title>Discord导出</title>
<style>body{font-family:'Segoe UI',sans-serif;background:#36393f;color:#dcddde;padding:20px}
.thread{background:#2f3136;margin:20px 0;border-radius:8px;overflow:hidden}
.thread-header{background:#5865f2;color:white;padding:15px;font-size:18px}
.message{padding:10px 15px;border-bottom:1px solid #40444b}
.author{color:#7289da;font-weight:bold}.time{color:#72767d;font-size:12px;margin-left:10px}
.content{margin-top:5px;white-space:pre-wrap}.link a{color:#00aff4}</style></head><body>"""
    for thread in threads_data:
        html += f'<div class="thread"><div class="thread-header">{thread["name"]} ({thread["created"]})</div>'
        for msg in thread["messages"]:
            html += f'<div class="message"><span class="author">{msg["author"]}</span><span class="time">{msg["time"]}</span>'
            html += f'<div class="content">{msg["content"]}</div>'
            html += f'<div class="link"><a href="{msg["link"]}" target="_blank">查看原消息</a></div></div>'
        html += '</div>'
    html += "</body></html>"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)


def do_export(urls, date_from, date_to, export_format):
    """后台执行导出任务"""
    global task_status
    try:
        all_threads_data = []
        total_messages = 0

        for i, url in enumerate(urls):
            task_status["progress"] = f"处理频道 {i+1}/{len(urls)}..."
            guild_id, channel_id = parse_discord_url(url)
            if not channel_id:
                continue

            channel_info = get_channel_info(channel_id)
            if not channel_info:
                continue

            channel_type = channel_info.get("type")
            channel_name = channel_info.get("name", "未知频道")

            if channel_type == 15:
                threads = get_all_threads(channel_id)
                task_status["progress"] = f"频道 {i+1}: 找到 {len(threads)} 个帖子"
                
                for j, thread in enumerate(threads):
                    thread_created = snowflake_to_datetime(thread["id"])
                    if date_from and thread_created < date_from:
                        continue
                    if date_to and thread_created > date_to:
                        continue

                    task_status["progress"] = f"频道 {i+1}: 处理帖子 {j+1}/{len(threads)}"
                    messages = get_channel_messages(thread["id"], date_from, date_to)
                    if messages:
                        messages.sort(key=lambda m: m["id"])
                        thread_data = {
                            "name": thread["name"],
                            "created": thread_created.strftime("%Y-%m-%d %H:%M"),
                            "messages": []
                        }
                        for msg in messages:
                            thread_data["messages"].append({
                                "author": msg.get("author", {}).get("username", "未知"),
                                "time": snowflake_to_datetime(msg["id"]).strftime("%Y-%m-%d %H:%M:%S"),
                                "content": msg.get("content", ""),
                                "attachments": "\n".join([a.get("url", "") for a in msg.get("attachments", [])]),
                                "link": f"https://discord.com/channels/{guild_id}/{thread['id']}/{msg['id']}"
                            })
                            total_messages += 1
                        all_threads_data.append(thread_data)
                    time.sleep(0.3)
            else:
                messages = get_channel_messages(channel_id, date_from, date_to)
                if messages:
                    messages.sort(key=lambda m: m["id"])
                    thread_data = {
                        "name": f"#{channel_name}",
                        "created": snowflake_to_datetime(channel_id).strftime("%Y-%m-%d %H:%M"),
                        "messages": []
                    }
                    for msg in messages:
                        thread_data["messages"].append({
                            "author": msg.get("author", {}).get("username", "未知"),
                            "time": snowflake_to_datetime(msg["id"]).strftime("%Y-%m-%d %H:%M:%S"),
                            "content": msg.get("content", ""),
                            "attachments": "\n".join([a.get("url", "") for a in msg.get("attachments", [])]),
                            "link": f"https://discord.com/channels/{guild_id}/{channel_id}/{msg['id']}"
                        })
                        total_messages += 1
                    all_threads_data.append(thread_data)

        if not all_threads_data:
            task_status["error"] = "没有找到符合条件的消息"
            return

        task_status["progress"] = "正在生成文件..."
        os.makedirs("exports", exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if export_format == "excel":
            filename = f"exports/Discord导出_{timestamp}.xlsx"
            export_to_excel(all_threads_data, filename)
        elif export_format == "txt":
            filename = f"exports/Discord导出_{timestamp}.txt"
            export_to_txt(all_threads_data, filename)
        else:
            filename = f"exports/Discord导出_{timestamp}.html"
            export_to_html(all_threads_data, filename)

        task_status["result"] = {"threads": len(all_threads_data), "messages": total_messages}
        task_status["filename"] = filename
        task_status["progress"] = "完成！"

    except Exception as e:
        task_status["error"] = str(e)
    finally:
        task_status["is_running"] = False
        if export_lock.locked():
            export_lock.release()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/check_token")
def check_token():
    return jsonify({"has_token": bool(BOT_TOKEN)})


@app.route("/api/set_token", methods=["POST"])
def set_token():
    global BOT_TOKEN
    data = request.json
    BOT_TOKEN = data.get("token", "")
    return jsonify({"success": True})


@app.route("/api/export", methods=["POST"])
def export():
    global BOT_TOKEN, task_status

    if not BOT_TOKEN:
        return jsonify({"error": "请先设置Bot Token"}), 400

    if not export_lock.acquire(blocking=False):
        return jsonify({
            "status": "busy",
            "message": "当前有任务正在执行，请稍后再试",
            "progress": task_status["progress"]
        }), 429

    try:
        data = request.json
        urls = data.get("urls", [])
        date_from = data.get("date_from")
        date_to = data.get("date_to")
        export_format = data.get("format", "excel")

        if not urls:
            export_lock.release()
            return jsonify({"error": "请输入至少一个频道链接"}), 400

        # 解析日期时间
        date_from_parsed = None
        date_to_parsed = None
        try:
            if date_from and date_from.strip():
                # 尝试解析带时间的格式 (datetime-local)
                try:
                    date_from_parsed = datetime.strptime(date_from.strip(), "%Y-%m-%dT%H:%M")
                except ValueError:
                    # 兼容旧的仅日期格式
                    date_from_parsed = datetime.strptime(date_from.strip(), "%Y-%m-%d")
            
            if date_to and date_to.strip():
                # 尝试解析带时间的格式 (datetime-local)
                try:
                    date_to_parsed = datetime.strptime(date_to.strip(), "%Y-%m-%dT%H:%M")
                except ValueError:
                    # 兼容旧的仅日期格式，自动设置为当天结束
                    date_to_parsed = datetime.strptime(date_to.strip(), "%Y-%m-%d")
                    date_to_parsed = date_to_parsed.replace(hour=23, minute=59, second=59)
        except Exception as e:
            export_lock.release()
            return jsonify({"error": f"日期时间格式错误: {str(e)}"}), 400

        # 重置状态
        task_id = str(uuid.uuid4())[:8]
        task_status["is_running"] = True
        task_status["task_id"] = task_id
        task_status["progress"] = "任务启动中..."
        task_status["result"] = None
        task_status["error"] = None
        task_status["filename"] = None

        # 启动后台线程
        thread = threading.Thread(target=do_export, args=(urls, date_from_parsed, date_to_parsed, export_format))
        thread.daemon = True
        thread.start()

        return jsonify({"status": "started", "task_id": task_id, "message": "导出任务已启动"})

    except Exception as e:
        if export_lock.locked():
            export_lock.release()
        return jsonify({"error": str(e)}), 500


@app.route("/api/status")
def get_status():
    return jsonify(task_status)


@app.route("/api/download/<path:filename>")
def download(filename):
    filepath = os.path.join(os.getcwd(), filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({"error": "文件不存在"}), 404


if __name__ == "__main__":
    os.makedirs("exports", exist_ok=True)
    app.run(debug=True, port=5000)
